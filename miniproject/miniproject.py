import csv
import pandas as pd
import logging
import openpyxl
import os
from datetime import date
from datetime import time






def getCallData(name):
	global logging
	try:
		
		wb= openpyxl.load_workbook(filename=name)
		logging.info(str(date.today())+": grabbing call data")
		
		name = name.replace(".xlsx","")
		items = name.split("_")
		month =items[-2]
		year= items[-1]
		
		months = {"january":1,"february":2,"march":3,"april":4,"may":5,"june":6,"july":7,"august":8,"september":9,"october":10,"november":11,"december":12}
		
		
		if month not in months:
			raise ValueError("Bad Name")
		
		sheet = wb.active
		tolerance =100
		#sheet.dimensions()
		
		sol=[]
		headers ={"Calls Offered":0,"Abandon after 30s":0,"FCR":0,"DSAT":0,"CSAT":0}
		found=False
		logging.info(str(date.today())+": Looking for data... ")
		for i in range(1,tolerance):
			for j in range(1,tolerance):
				cell = sheet.cell(row=i,column=j)
				
				wk=str(cell.value).split("-") #de-format the cell
				if(len(wk ) >1):
					entry =(wk[0],int(wk[1])) #grab date
					if(year ==entry[0] and months[month]== entry[1]):
						for k in range(1,6):
							sol.append(sheet.cell(row=i,column=j+k).value) #get the 6 data values
						found = True
						logging.info(str(date.today())+": Found entry at "+str(i)+","+str(j))
						
						break #found data, we're done
				elif(len(wk) ==1):
					if(wk[0].strip() in headers):
						headers[wk[0].strip()]=j #grab/verify the headers
						logging.info(str(date.today())+": Found Header at "+str(i)+","+str(j))
			if found:
				logging.info(str(date.today())+": Found All required entries")
				break
		if not found:
			logging.error(str(date.today())+": Exausted search within tolerance value")
			print("Data not found")
			return
		
		headers = dict(sorted(headers.items(), key=lambda item: item[1])) ##sort the headers list
		head_l = list(headers.keys()) #makes headers indexable in order they were retireved in
		logging.info(str(date.today())+": Sorted Headers list for display")
		logging.info(str(date.today())+": Displaying data found in "+name)
		print("Call data output for",name)
		print(head_l[0],": ",sol[0])
		logging.info(str(date.today())+": Displayed item 1: "+str(head_l[0])+": "+str(sol[0]))
		print(head_l[1],": ",sol[1]*100,"%")
		logging.info(str(date.today())+": Displayed item 2: "+str(head_l[1])+": "+str(sol[1])+"%")
		print(head_l[2],": ",sol[2]*100,"%")
		logging.info(str(date.today())+": Displayed item 3: "+str(head_l[2])+": "+str(sol[2])+"%")
		print(head_l[3],": ",sol[3]*100,"%")
		logging.info(str(date.today())+": Displayed item 4: "+str(head_l[3])+": "+str(sol[3])+"%")
		print(head_l[4],": ",sol[4]*100,"%")
		logging.info(str(date.today())+": Displayed item 5: "+str(head_l[4])+": "+str(sol[4]) +"%")
		
		
		
		#print("---------")
		
		
		logging.info(str(date.today())+": Program Finished")
		
		
			
			
	except FileNotFoundError as e:
		print("Could not find that file")
		
		logging.error(str(date.today())+": Tried to open nonexistant file with name: "+name)
	except openpyxl.utils.exceptions.InvalidFileException as e:
		print("Invalid File")
		
		logging.error(str(date.today())+": Tried to open invalid file with name: "+name)
		
		
		
		
def getVOCData(name):
	global logging
	
	try:
		wb= openpyxl.load_workbook(filename=name)
		logging.info(str(date.today())+": grabbing VOC data from "+name)
		
		name = name.replace(".xlsx","")
		items = name.split("_")
		month =items[-2]
		year= items[-1]
		
		months = {"january":1,"february":2,"march":3,"april":4,"may":5,"june":6,"july":7,"august":8,"september":9,"october":10,"november":11,"december":12}
		
		sheet = wb["VOC Rolling MoM"]
		#sheet=wb["Test"]
		
		headers ={"promoters":0,"passives":0,"dectractors":0}
		
		tolerance =100
		col =0
		foundCol = False
		rows=[]
		#print(month)
		for i in range(1,tolerance):
			for j in range(1,tolerance):
				cell = sheet.cell(row=i,column=j)
				
				wk=str(cell.value).split("-")
				#print(wk)
				if(not foundCol):
					if((len(wk)==1 and wk[0].lower() in months) or (len(wk) >1 and int(wk[1].lower()) ==months[month])):#should find first instance of month
						col = j
						foundCol = True
				wk = str(cell.value).split(" ")
				if(wk[0].lower() in headers):
					#rows.append(i)
					headers[wk[0].lower()] = i
					
		sol=[]
		#print(headers)
		for key in headers:
			sol.append(sheet.cell(row=headers[key],column=col).value)
			
			
		
		print("VOC data for",name)
		i=0
		for key in headers:
			print(key.capitalize(),": ",determine(key,sol[i]))
			logging.info(str(date.today())+" Entry for "+key+" is "+determine(key,sol[i]))
			i+=1
			
		#print("----------")
				
					
		
		#print(sheet)
		
		
		
		
		
		
		
		
		
		
		
	except FileNotFoundError as e:
		print("Could not find that file")
		
		logging.error(str(date.today())+": Tried to open nonexistant file with name: "+name)
	

def determine(head,value):
	if head == "promoters":
		if value >= 200:
			return "good"
		else:
			return "bad"
	elif head =="passives":
		if value >=100:
			return "good"
		else:
			return "bad"
	elif head=="dectractors":
		if value>=100:
			return "good"
		else:
			return "bad"
	else:
		return "bad"
		




def main():
	logging.basicConfig(filename="log.log",level=logging.DEBUG)
	
	logging.info(str(date.today())+": Started Program")
	validFiles =["xlsx","csv"]
	workFiles =[]
	#print(os.listdir())
	
	
	#print(os.getcwd())
	
	
	here = os.getcwd()
	arch = os.path.join(here,r'archive')
	if  not os.path.exists(arch):
		#print("here")
		os.makedirs(arch)
		
	error = os.path.join(here,r'error')
	if not os.path.exists(error):
		os.makedirs(error)
	
	
	
	
	for file in os.listdir():
		
		if file.split(".")[-1] in validFiles:
			workFiles.append(file)
			
	try:		
		f = open("filelist.lst","r")
		lines = f.read().splitlines()
	except FileNotFoundError:
		f = open("filelist.lst","w")
		lines =[]
	
	f.close()
	
	#print(lines)
	#print(workFiles)
	
	#ip = input("Enter name of files: ")
	
	#check all xlsx files in directory, don't use inputs from user
	#year is not guaanteed for second set of data
	
	
	#names=ip.split(" ")
	#print(names)
	f=open("filelist.lst","a")
	for name in workFiles:
		try:
			if name not in lines:
				getCallData(name)
				print("-----")
				getVOCData(name)
				print("----------")
				f.write(name+'\n')
				os.replace(os.path.join(here,name),os.path.join(arch,name))
			else:
				print(name,"was already processed!")
				logging.warning(str(date.today())+" : "+ name+" was already processed!")
		except ValueError as e:
			print("Poorly Formatted File")
			logging.error(str(date.today())+": Tried to open badly-formatted file with name: "+name)
			os.replace(os.path.join(here,name),os.path.join(error,name))
			
			
		except KeyError as e:
			print("Incomplete Data File")
			logging.error(str(date.today())+": Tried to open file with incomeplete data: "+name)
			os.replace(os.path.join(here,name),os.path.join(error,name))
			
			
	print("Check log.log for more details")
	
	
	
	
	
	
	
	
main()


